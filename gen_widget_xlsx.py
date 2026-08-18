#!/usr/bin/env python3
"""生成 widget 模型用量公式表 (xlsx)"""
import json, urllib.request, zipfile, re, os, sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import formula_registry as fr

# 尝试 openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

API = "http://127.0.0.1:8765/api/state"
OUT = os.path.join(os.path.expanduser("~"), "Desktop", "widget-formula.xlsx")
if os.path.exists(OUT):
    base, ext = os.path.splitext(OUT)
    OUT = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"

# ---------- 拉取 API ----------
def fetch_state():
    r = urllib.request.urlopen(API, timeout=15)
    return json.loads(r.read().decode("utf-8"))

# ---------- 读取桌面 widget.xlsx 模板（可选） ----------
def read_widget_template(path):
    """解析桌面 widget.xlsx，返回表头和数据行结构。"""
    z = zipfile.ZipFile(path)
    shared = []
    try:
        sx = z.read("xl/sharedStrings.xml").decode("utf-8")
        for si in re.findall(r"<si>(.*?)</si>", sx, re.DOTALL):
            txt = "".join(re.findall(r"<t[^>]*>(.*?)</t>", si, re.DOTALL))
            shared.append(txt)
    except KeyError:
        pass
    import xml.etree.ElementTree as ET
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    for name in z.namelist():
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
            xml = z.read(name).decode("utf-8")
            root = ET.fromstring(xml)
            rows = root.findall(".//m:sheetData/m:row", ns)
            data = []
            for r in rows:
                rn = r.get("r")
                row = {}
                for c in r.findall("m:c", ns):
                    ref = c.get("r")
                    col = re.match(r"([A-Z]+)", ref).group(1)
                    t = c.get("t")
                    v = c.find("m:v", ns)
                    val = v.text if v is not None else ""
                    if t == "s":
                        try:
                            val = shared[int(val)]
                        except:
                            pass
                    row[col] = val
                data.append((rn, row))
            return data
    return []

# ---------- 构建模型数据 ----------
def build_model_rows(state):
    stats = state.get("stats", [])
    # 按 model 去重，source=go 优先（官方权威），gateway 仅补充
    seen = {}
    for m in stats:
        if m.get("group") != "go":
            continue
        model = m.get("model")
        src = m.get("source", "")
        if model not in seen:
            seen[model] = m
        else:
            if src == "go" and seen[model].get("source") != "go":
                seen[model] = m
    models = sorted(seen.values(), key=lambda x: x.get("model", ""))
    return models

# ---------- 生成 xlsx ----------
def generate_xlsx(models, state, template_rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模型用量数据"

    # 样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # ---------- Sheet1: 模型用量数据 ----------
    # 表头
    headers = [
        "模型", "来源",
        "今天token", "7天token", "30天token", "全部token",
        "今天次数", "7天次数", "30次数", "全部次数",
        "今日金额", "七天金额", "三十天金额",
        "model_quota", "req_lim[2]", "tokens/次", "est_req_cost", "est_tok_cost",
        "model_used", "model_remain",
        "本地gateway补充cost", "备注"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # 数据行
    for m in models:
        model = m.get("model", "")
        src = m.get("source", "")
        tok_s = (m.get("tokens_in_s") or 0) + (m.get("tokens_out_s") or 0) + (m.get("tokens_cache_s") or 0)
        tok_w = (m.get("tokens_in_w") or 0) + (m.get("tokens_out_w") or 0) + (m.get("tokens_cache_w") or 0)
        tok_m = (m.get("tokens_in_m") or 0) + (m.get("tokens_out_m") or 0) + (m.get("tokens_cache_m") or 0)
        tok_all = (m.get("tokens_in") or 0) + (m.get("tokens_out") or 0) + (m.get("tokens_cache") or 0)

        gateway_m = None
        for mm in state.get("stats", []):
            if mm.get("model") == model and mm.get("source") == "gateway":
                gateway_m = mm
                break

        def F(val, src, expr):
            if val is None or val == "":
                return val
            return f"{val}\n{src}\n{expr}"

        row = [
            model,
            src,
            F(tok_s, "官方 usage_records", "tokens_in_s + tokens_out_s + tokens_cache_s"),
            F(tok_w, "官方 usage_records", "tokens_in_w + tokens_out_w + tokens_cache_w"),
            F(tok_m, "官方 usage_records", "tokens_in_m + tokens_out_m + tokens_cache_m"),
            F(tok_all, "官方 usage_records", "tokens_in + tokens_out + tokens_cache"),
            F(m.get("count_s") or 0, "官方 usage_records", "count_s"),
            F(m.get("count_w") or 0, "官方 usage_records", "count_w"),
            F(m.get("count_m") or 0, "官方 usage_records", "count_m"),
            F(m.get("count_total") or 0, "官方 usage_records", "count_total"),
            F(m.get("cost_s") or 0.0, "官方 API 返回原始值", "cost_s"),
            F(m.get("cost_w") or 0.0, "官方 API 返回原始值", "cost_w"),
            F(m.get("cost_m") or 0.0, "官方 API 返回原始值", "cost_m"),
            F(m.get("cost_total") or 0.0, "官方 API 返回原始值", "cost_total"),
            F(m.get("model_quota"), "云端 params.model_quotas", "MODEL_QUOTAS[m]"),
            F(m.get("req_lim", [None])[2] if m.get("req_lim") else None, "云端 params.req_limits", "REQ_LIMITS[m][2]"),
            F(state.get("params", {}).get("tokens_per_req", {}).get(model) or m.get("tokens_per_req"), "云端 params.tokens_per_req", "TOKENS_PER_REQ[m]"),
            F(m.get("est_req_cost"), "云端 params.model_quotas + req_limits", "model_quota / req_limits[2]"),
            F(m.get("est_tok_cost"), "est_req_cost + 云端 params.tokens_per_req", "est_req_cost / tokens_per_req"),
            F(m.get("model_used"), "官方 cost_map 权威 / 本地聚合", "cost_map[m] (go优先)"),
            F(m.get("model_remain"), "model_quota + model_used", "max(0, model_quota - model_used)"),
            F(gateway_m.get("cost_total") if gateway_m else None, "本地 gateway 补充", "gateway source cost_total"),
            "go(官方) + gateway(本地补充)" if gateway_m else "",
        ]
        ws.append(row)

    # 供应商汇总行
    ws.append(["--- 供应商汇总 ---"] + [""] * (len(headers) - 1))
    suppliers = state.get("suppliers", {})
    if isinstance(suppliers, dict):
        for src, s in sorted(suppliers.items()):
            tok_all = s.get("tokens") or 0

            def F(val, src, expr):
                if val is None or val == "":
                    return val
                return f"{val}\n{src}\n{expr}"

            row = [
                f"[{src}] 供应商",
                src,
                "", "", "", F(tok_all, "供应商聚合", "tokens"),
                "", "", "", F(s.get("count") or 0, "供应商聚合", "count"),
                "", "", "", F(s.get("cost") or 0.0, "供应商聚合", "cost"),
                "", "", "", "", "",
                F(s.get("cost") or 0.0, "供应商聚合", "cost"),
                ""
            ]
            ws.append(row)

    # 自动列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 50
        for cell in row:
            if isinstance(cell.value, str) and "\n" in cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ---------- Sheet2: 公式说明 ----------
    ws2 = wb.create_sheet(title="公式说明")
    ws2.append(["公式ID", "公式名称", "来源", "计算方式", "代码位置", "备注"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row in fr.FORMULA_TABLE():
        ws2.append([
            row.get("id", ""),
            row.get("display", ""),
            row.get("source", ""),
            row.get("expr", ""),
            row.get("used_by", ""),
            row.get("params", ""),
        ])

    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws2.column_dimensions[column].width = adjusted_width

    # 保存
    wb.save(OUT)
    print(f"已生成: {OUT}")

# ---------- 主流程 ----------
if __name__ == "__main__":
    print("拉取 API 数据...")
    state = fetch_state()
    models = build_model_rows(state)
    print(f"模型数: {len(models)}")
    # 读取桌面模板（可选参考）
    template_path = os.path.join(os.path.expanduser("~"), "Desktop", "widget.xlsx")
    template_rows = None
    if os.path.exists(template_path):
        template_rows = read_widget_template(template_path)
        print(f"读取模板: {len(template_rows)} 行")
    generate_xlsx(models, state, template_rows)
